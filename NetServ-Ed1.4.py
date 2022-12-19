#!/usr/bin/env python
# coding: utf-8

# In[2]:


import pandas as pd
import xlsxwriter
import openpyxl 
import re
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule

#list of rates required for each site
ListOfRates = ['1GbE' , '10GbE' , '100GBE', 'FC400' , 'FC800' , 'ODU2', 'STM-1' , 'STM-16' , 'STM-4' , 'STM-64' , 'DSR']
ListOfSites = ['SGS_JME' , 'SGC_JGE' , 'SGC_MTA', 'SGS_JFL' , 'NGA_PCE' , 'SGC_MDP' , 'NGA_PPQ' , 'NGA_PRO' , 'SGC_Teraco' ,
                'NGA_PSI' , 'DFA_EMH' , 'DFA_ER' , 'DFA_MTZ' , 'KZN_RV' , 'KZN_DMO' , 'KZN_DNE' , 'KZN_DTA' , 'WC_FSDC' ,
                'WC_Terraco' , 'WC_YZN' , 'WC_CTE' , 'EAS_EL-TELKOM' , 'EAS_EL' , 'EL-Telkom-1' ,'NLD10_EL' ,
                'EC_EL_Telkom' , 'EAS_EL' , 'CEN_BES' , 'NLD8_POL_TEL' ,'NLD8_POL-TEL', 'NLD7_POL-CUBE' , 'NL_Telkom' , 'NL_' ]
    
#location of source path
PathOrigin = r"D:\PROJECTS\SA\Migration-prepartion-2\SNC-SharedRisk-Report.csv"

#to read the excel sheet
dataSource = pd.read_csv (PathOrigin)

#to arrange columns as new required
df = pd.DataFrame(dataSource, columns = [ 'Rate' , 'Protection' , 'Name' , 'Servers' , 'Service Trails' , 'Service OTS' ,
                                  'Protetion Trails' , 'Protection OTS' ] )

# a new list of sites with '-' in naming 
ListOfAdjustedSites = [w.replace('_', '-') for w in ListOfSites]
z=0 ; z=int(z)
# a loop for replacing all sites name
while z < len(ListOfAdjustedSites) :    
    df['Name']=df['Name'].apply(lambda x:str(x.replace(ListOfAdjustedSites[z],ListOfSites[z])))
    z += 1

# intiatining values used for filtering rates coulumn
x=0 ; x=int(x)

# a loop for repeating all sites
while x < len(ListOfSites) :
        
    # intiatining values used for filtering sites coulumn
    y=0 ; y=int(y)
    #iniatializating readed datafram
    data = []
   
    # a loop for repeating all rates
    while y < len(ListOfRates) :
        
        #to filer on site name and rate plus separting routes each in a line 
        sh = df[df['Rate'].str.endswith(ListOfRates[y]) & df['Name'].str.contains(ListOfSites[x])]
        str_cols = ['Service Trails' ,'Service OTS' ,'Protetion Trails' ,'Protection OTS' ]
        sh[str_cols] = sh[str_cols].replace(',','\n', regex=True)
        sh['Servers']=sh['Servers'].apply(lambda x:str(x.replace("['","").replace("']", "").replace(',','\n').replace("'", "")))        
                                                                       
        #condition to exclude empty filtered sheets
        if len(sh.index) > 0 :
            
            #location of output files
            PathDestinatioin = (r'D:\PROJECTS\SA\Migration-prepartion-2\Output\XC-report - ' + ListOfSites[x] 
                                + '.xlsx')
            
            # calcutaling the width for some columns
            width= sh['Name'].str.len().max()
            width2 = len('protection') + 2
            width3 =(len(sh.loc[sh.Servers != '\n', 'Servers'].max()))
            width4 = 66
            width5 = 122
                       
            #condition to add tabs in same excel sheet
            if len(data) == 0 :
                
                # starting xlsxwriter
                writer = pd.ExcelWriter(PathDestinatioin, engine='xlsxwriter' , mode='w')
                #save the datafram
                sh.to_excel ( writer , sheet_name = ListOfRates[y] , index=False , header=True)
                #call saved book & sheet to Access the XlsxWriter workbook and worksheet objects from the dataframe
                wb = writer.book
                ws = writer.sheets[ListOfRates[y]]
                
                # header required format
                header_format = wb.add_format({'bold': True,'text_wrap': False,'valign': 'top','fg_color': '#D7E4BC',
                                                     'border': 1})
                                
                #apply this format for all colums header
                for col_num, value in enumerate(sh.columns.values):
                    ws.write(0, col_num, value, header_format)
                
                # freezing frist three headers 
                ws.freeze_panes(1, 3)
                
                # bolding Font for a spcific column
                bold_fmt = wb.add_format({'bold': True,'valign': 'top'})
                                
                #adjusting the width for some columns
                ws.set_column(2, 2, width + 3, bold_fmt)
                ws.set_column(1, 1, width2 - 1)
                
                if width3 > 60 : ws.set_column(3, 3, width3 + 3)
                elif width3 < 60 : ws.set_column(3, 3, width3 + 60)
                else : ws.set_column(3, 3, width3 + 60)
                    
                ws.set_column(4, 4, width4 )
                ws.set_column(5, 5, width5 )
                ws.set_column(6, 6, width4 )
                ws.set_column(7, 7, width4 +44 )
                
                # adjusting zooming for 1st tab
                ws.set_zoom(75)
                
                #applying Warp text for cloumn D to H
                wrap_format = wb.add_format({'text_wrap': True,'valign': 'top'})
                ws.set_column('D:H', 70, wrap_format)
                
                writer.save() 
                writer.close()
                
                #reading the stored sheet again to use it in next update (adding extra tabs)
                data = pd.read_excel(PathDestinatioin)
            else :
                
                # starting openXLpy
                writer = pd.ExcelWriter(PathDestinatioin, engine='openpyxl' , mode='a')                
                sh.to_excel( writer, sheet_name = ListOfRates[y] , index=False , header=True)
                
                #call the sheet & work book for formatting
                wb = writer.book
                ws = writer.sheets[ListOfRates[y]]
                
                #adjusting the width for some columns
                ws.column_dimensions["C"].width = width+3                
                ws.column_dimensions["B"].width = width2 - 1
                
                if width3 > 60 : ws.column_dimensions["D"].width = width3+3
                elif width3 < 60 : ws.column_dimensions["D"].width = width3 +60
                else : ws.column_dimensions["D"].width = width3+60
                
                ws.column_dimensions["E"].width = width4
                ws.column_dimensions["F"].width = width5
                ws.column_dimensions["G"].width = width4
                ws.column_dimensions["H"].width = width4 + 44
               
                # freezing frist three headers
                ws.freeze_panes = ws['D2']
                                                                
                #appling coditioninatal formating color
                color_Fill = PatternFill(bgColor="D7E4BC")                
                ws.conditional_formatting.add('A1:H1',FormulaRule(formula=['ISBLANK(I1)'], stopIfTrue=True, fill=color_Fill))
                
                # bolding Font for a spcific column
                for i in range(len(sh.index)) :
                    m="c"
                    ws[m + str(i+2)].font = Font(bold=True)
                
                # adjusting zooming for remaining tabs
                for ws in wb.worksheets:
                    ws.sheet_view.zoomScale = 75
                    
                #applying Warp text for all cloumn 
                for rows in ws.iter_rows():
                    for cell in rows:
                        cell.alignment = Alignment(wrap_text=True)
                
                writer.save()
                writer.close()
               
        y += 1

    x += 1


# In[ ]:




