import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# Define lists of rates and sites
rates = ['1GbE' , '10GbE' , '100GBE', 'FC400' , 'FC800' , 'ODU2', 'STM-1' , 'STM-16' , 'STM-4' , 'STM-64' , 'DSR']
sites = ['SGS_JME' , 'SGC_JGE' , 'SGC_MTA', 'SGS_JFL' , 'NGA_PCE' , 'SGC_MDP' , 'NGA_PPQ' , 'NGA_PRO' , 'SGC_Teraco' ,
                'NGA_PSI' , 'DFA_EMH' , 'DFA_ER' , 'DFA_MTZ' , 'KZN_RV' , 'KZN_DMO' , 'KZN_DNE' , 'KZN_DTA' , 'WC_FSDC' ,
                'WC_Terraco' , 'WC_YZN' , 'WC_CTE' , 'EAS_EL-TELKOM' , 'EAS_EL' , 'EL-Telkom-1' ,'NLD10_EL' ,
                'EC_EL_Telkom' , 'EAS_EL' , 'CEN_BES' , 'NLD8_POL_TEL' ,'NLD8_POL-TEL', 'NLD7_POL-CUBE' , 'NL_Telkom' , 'NL_']

# Read data from CSV using pandas
df = pd.read_csv("SNC-SharedRisk-Report.csv")

# Define formatting styles
header_format = {
    "bold": True,
    "text_wrap": False,
    "valign": "top",
    "fg_color": "#D7E4BC",
    "border": 1,
}
bold_format = {"bold": True, "valign": "top"}
wrap_format = {"text_wrap": True, "valign": "top"}

# Function to process data and generate report for a site and rate
def process_site_rate(site, rate):
    filtered_df = df[(df["Rate"] == rate) & df["Name"].str.contains(site)]

    # Replace commas with newlines efficiently
    filtered_df["Service Trails"] = filtered_df["Service Trails"].str.replace(
        ",", "\n", regex=True
    )
    filtered_df[["Service OTS", "Protetion Trails", "Protection OTS"]] = (
        filtered_df[["Service OTS", "Protetion Trails", "Protection OTS"]].replace(
            ",", "\n", regex=True
        )
    )

    # Replace brackets and commas in "Servers" efficiently
    filtered_df["Servers"] = filtered_df["Servers"].apply(
        lambda x: x.replace("['", "").replace("']", "").replace(",", "\n").replace("'", "")
    )

    # Create Excel workbook and sheet if data exists
    if not filtered_df.empty:
        wb = Workbook()
        ws = wb.active
        ws.append(filtered_df.columns.tolist())

        # Format headers and data
        for col_num, value in enumerate(filtered_df.columns.values):
            ws.cell(row=1, column=col_num + 1).value = value
            ws.cell(row=1, column=col_num + 1).font = Font(**header_format)

        for row_num in range(2, len(filtered_df) + 2):
            for col_num in range(1, len(filtered_df.columns) + 1):
                ws.cell(row=row_num, column=col_num).value = filtered_df.iloc[row_num - 2, col_num - 1]
                ws.cell(row=row_num, column=col_num).alignment = Alignment(**wrap_format)

        # Additional formatting
        ws.freeze_panes = ws["D2"]
        ws.row_dimensions[1].height = 35

        column_widths = {
            "B": len("protection") + 2,
            "C": max(len(x) for x in filtered_df["Servers"]) + 3,
            "D": 66,
            "E": 122,
            "F": 66,
            "G": 110,
            "H": 66,
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        for i in range(len(filtered_df)):
            ws.cell(row=i + 2, column=1).font = Font(**bold_format)

        # Save the Excel file
        wb.save(f"XC-report - {site} - {rate}.xlsx")


# Process data for each site and rate
for site in sites:
    for rate in rates:
        process_site_rate(site, rate)

print("Reports generated successfully!")